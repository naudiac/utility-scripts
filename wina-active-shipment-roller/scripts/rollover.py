import sys
import openpyxl
from datetime import datetime

def rollover_shipments(filepath):
    print(f"Loading {filepath} (data_only=True) to map rows for deletion...")
    wb_data = openpyxl.load_workbook(filepath, data_only=True)
    sheet_data = wb_data['GWS Report template']

    status_col = None
    report_date_col = None
    for cell in sheet_data[1]:
        val = str(cell.value).strip().lower() if cell.value else ""
        if val == 'shipment status':
            status_col = cell.column
        elif val == 'shipment report date':
            report_date_col = cell.column
            
    if not status_col or not report_date_col:
        print("Error: Could not find 'Shipment Status' or 'Shipment report date' columns.")
        wb_data.close()
        return

    rows_to_delete = []
    # Start from bottom to top
    for row_idx in range(sheet_data.max_row, 1, -1):
        status_cell = sheet_data.cell(row=row_idx, column=status_col).value
        status_val = str(status_cell).strip().lower() if status_cell else ""
        if status_val in ['arrived', 'cancelled']:
            rows_to_delete.append(row_idx)
            
    wb_data.close()
    
    print(f"Mapped {len(rows_to_delete)} rows for deletion. Loading formula-preserved master workbook...")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['GWS Report template']
    
    # Delete those rows
    for row_idx in rows_to_delete:
        sheet.delete_rows(row_idx, 1)
        
    # Update dates
    today_date = datetime.now()
    for row_idx in range(2, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=report_date_col).value = today_date
        
    output_path = filepath.replace('.xlsx', '_ROLLED.xlsx')
    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print("Rollover complete!")

if __name__ == "__main__":
    rollover_shipments(sys.argv[1])
