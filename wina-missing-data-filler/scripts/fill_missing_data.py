import openpyxl
from openpyxl.styles import PatternFill, Font
import datetime
import pyodbc
import os
import json
import re
import argparse
import sys

# ==========================================
# CONFIGURATION
# ==========================================
NULL_VARIATIONS = ['', 'null', 'na', 'n/a', 'not available', 'none', 'tbd']

YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
CLEAR_FILL = PatternFill(fill_type=None)
RED_FONT = Font(color='FFFF0000')      # Asia
BLUE_FONT = Font(color='FF0000FF')     # America
ORANGE_FONT = Font(color='FFFFA500')   # EMEA
BLACK_FONT = Font(color='FF000000')    # Default

DELIVERY_FIELDS = ['Discharge Date', 'FRN', 'Gate Out Date', 'Actual delivery to door date', 'Empty return Date']
CHECK_FIELDS = ['Container No.', 'No. of Units (Pallets/Packages)', 'UOM', 'Total CMB', 'Total Weight in KG', 'Invoice Number', 'Container Type', 'Requested pick up date', 'Empty pick-up Date', 'Actual Pick up Date', 'Gate In Date', 'ATD Date', 'ATA Date'] + DELIVERY_FIELDS

HIGHLIGHT_FIELDS = [
    'Actual Pick up Date', 
    'ATD Date', 
    'ATA Date', 
    'Actual delivery to door date', 
    'Container No.', 
    'No. of Units (Pallets/Packages)', 
    'UOM', 
    'Total CMB', 
    'Total Weight in KG', 
    'Container Type'
]

DATE_COLUMNS = [
    'Requested pick up date', 'Empty pick-up Date', 'PLANNED PICK UP DATE', 'Revised planned pick up date', 
    'Actual Pick up Date', 'Gate In Date', 'ORIGINAL ETD DATE', 'Revised ETD Date', 
    'ATD Date', 'ORIGINAL ETA DATE', 'Revised ETA date', 'ATA Date', 
    'Discharge Date', 'FRN', 'Gate Out Date', 'PLANNED DELIVERY TO DOOR DATE', 
    'Revised planned delivery to door date', 'Actual delivery to door date', 
    'Empty return Date'
]


# ==========================================
# DATABASE HELPER
# ==========================================
def clean_value(val):
    if val is None:
        return ''
    return str(val).strip().lower()

def get_cw_data(file_num, missing_fields):
    results = {}
    server = os.environ.get('CW_DB_SERVER')
    database = os.environ.get('CW_DB_NAME')
    username = os.environ.get('CW_DB_USER')
    password = os.environ.get('CW_DB_PASS')
    
    if not all([server, database, username, password]):
        return {}

    conn_str = f"DRIVER={{ODBC Driver 11 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
    
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        if 'Owner' in missing_fields:
            cursor.execute('''
                SELECT TOP 1 GS.GS_GivenName, GS.GS_Surname 
                FROM JobShipment JS 
                LEFT JOIN GlbStaff GS ON GS.GS_Code = JS.JS_SystemLastEditUser
                WHERE JS.JS_UniqueConsignRef = ?
            ''', file_num)
            row = cursor.fetchone()
            if row and row[0] and row[1]:
                results['Owner'] = f"{row[0]} {row[1]}"
                
        # Simulate fetching Invoice Number since we don't have the exact join configured yet
        # In full production, this would query JobComInvoiceHeader
        if 'Invoice Number' in missing_fields:
            pass # results['Invoice Number'] = 'some_val'
            
        if 'Requested pick up date' in missing_fields:
            pass # results['Requested pick up date'] = 'some_val'

    except Exception as e:
        print(f"DB Error for {file_num}: {e}")
    finally:
        if 'conn' in locals() and conn:
            conn.close()
            
    return results

def is_human_name(name):
    if not name or not isinstance(name, str): return False
    name_lower = name.lower()
    invalid_keywords = ['team', 'admin', 'system', 'import', 'export']
    if any(kw in name_lower for kw in invalid_keywords):
        return False
    return ' ' in name.strip()

def get_date_val(cell):
    if not cell or cell.value is None:
        return None
    if isinstance(cell.value, datetime.datetime):
        return cell.value
    return None

def is_overdue(planned_date, revised_date, actual_val, today):
    if actual_val is not None and str(actual_val).strip() != '':
        return False
    if revised_date is not None:
        return today > revised_date
    if planned_date is not None:
        return today > planned_date
    return False

def normalize_header(h):
    if not h: return ''
    return re.sub(r'\s+', ' ', str(h)).strip().lower()

# ==========================================
# MAIN SCRIPT
# ==========================================
def main():
    parser = argparse.ArgumentParser(description='Fill missing CW data in GWS Report.')
    parser.add_argument('input_file', help='Path to the GWS Report Excel file')
    args = parser.parse_args()
    
    excel_path = args.input_file
    output_path = excel_path.replace('.xlsx', '_UPDATED.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        sys.exit(1)

    print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb['GWS Report template']
    
    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[normalize_header(cell.value)] = cell.column
            
    today = datetime.datetime.now()
    
    def get_col(name):
        return headers.get(normalize_header(name), 0)
    
    try:
        col_file = get_col('File #')
        col_owner = get_col('Owner')
        col_status = get_col('Shipment Status')
        col_bol = get_col('BOL NO')
        col_ms_status = get_col('MILESTONE STATUS')
        
        col_atd = get_col('ATD Date')
        col_orig_etd = get_col('ORIGINAL ETD DATE')
        col_rev_etd = get_col('Revised ETD Date')
        
        col_ata = get_col('ATA Date')
        col_orig_eta = get_col('ORIGINAL ETA DATE')
        col_rev_eta = get_col('Revised ETA date')
        
        col_act_deliv = get_col('Actual delivery to door date')
        col_plan_deliv = get_col('PLANNED DELIVERY TO DOOR DATE')
        col_rev_deliv = get_col('Revised planned delivery to door date')
        
        col_origin_reg = get_col('Origin Region')
        col_origin_cntry = get_col('Origin Country')
        col_dest_reg = get_col('Destination Region')
        col_invoice = get_col('Invoice Number')
        col_service = get_col('SERVICE TYPE')
        
        if not col_file or not col_rev_etd:
            raise KeyError(f"Missing essential columns! col_file={col_file}, col_rev_etd={col_rev_etd}")
            
    except KeyError as e:
        print(f"Missing essential column header: {e}")
        return

    print("Processing rows...")
    for row in range(2, sheet.max_row + 1):
        file_num = sheet.cell(row=row, column=col_file).value
        if not file_num:
            continue
            
        ship_status = sheet.cell(row=row, column=col_status).value
        if ship_status == 'Cancelled':
            continue
            
        # 1. OWNER LOGIC
        owner_cell = sheet.cell(row=row, column=col_owner)
        owner_val = str(owner_cell.value).strip() if owner_cell.value else ""
        if not owner_val or owner_val.lower() == 'need owner':
            cw_data = get_cw_data(file_num, ['Owner'])
            if 'Owner' in cw_data and is_human_name(cw_data['Owner']):
                owner_cell.value = cw_data['Owner']
            else:
                owner_cell.value = "Need Owner"
                
        # INVOICE NUMBER LOGIC
        if col_invoice:
            inv_cell = sheet.cell(row=row, column=col_invoice)
            inv_val = clean_value(inv_cell.value)
            if inv_val in NULL_VARIATIONS:
                cw_data = get_cw_data(file_num, ['Invoice Number'])
                cw_inv = clean_value(cw_data.get('Invoice Number'))
                if cw_inv in NULL_VARIATIONS:
                    inv_cell.value = "NOT AVAILABLE"
                else:
                    inv_cell.value = cw_data['Invoice Number']
        
        # 2. DATE HARMONIZATION
        atd_val = sheet.cell(row=row, column=col_atd).value
        if atd_val:
            orig_etd = sheet.cell(row=row, column=col_orig_etd).value
            if atd_val != orig_etd:
                sheet.cell(row=row, column=col_rev_etd).value = atd_val
                
        ata_val = sheet.cell(row=row, column=col_ata).value
        if ata_val:
            orig_eta = sheet.cell(row=row, column=col_orig_eta).value
            if ata_val != orig_eta:
                sheet.cell(row=row, column=col_rev_eta).value = ata_val
                
        act_deliv_val = sheet.cell(row=row, column=col_act_deliv).value
        if act_deliv_val:
            plan_deliv = sheet.cell(row=row, column=col_plan_deliv).value
            if act_deliv_val != plan_deliv:
                sheet.cell(row=row, column=col_rev_deliv).value = act_deliv_val
        
        # 2.5 FORMATTING CLEANUP
        for col_name in DATE_COLUMNS:
            col_idx = get_col(col_name)
            if col_idx:
                c = sheet.cell(row=row, column=col_idx)
                c.fill = CLEAR_FILL
                c.font = BLACK_FONT

        # 3. OVERDUE LOGIC & HIGHLIGHTING
        service_type_cell = sheet.cell(row=row, column=col_service).value if col_service else None
        service_type = str(service_type_cell).strip().upper() if service_type_cell else ""
        
        has_pickup = service_type.startswith('D') if service_type else True
        has_delivery = service_type.endswith('D') if service_type else True
        
        overdue_highlight_fields = []
        is_delivery_only_overdue = True
        
        for field in CHECK_FIELDS:
            col_idx = get_col(field)
            if not col_idx:
                continue
                
            cell = sheet.cell(row=row, column=col_idx)
            
            # Skip pickup fields if no pickup service (clear any existing yellow highlight)
            if not has_pickup and ('pick up' in field.lower() or 'pickup' in field.lower() or 'pick-up' in field.lower()):
                cell.fill = CLEAR_FILL
                continue
                
            # Skip delivery fields if no delivery service (clear any existing yellow highlight)
            if not has_delivery and 'delivery' in field.lower():
                cell.fill = CLEAR_FILL
                continue
                
            val = cell.value
            field_overdue = False
            
            if val is None or clean_value(val) in NULL_VARIATIONS:
                planned_date = None
                revised_date = None
                
                if 'pickup' in field.lower():
                    planned_date = get_date_val(sheet.cell(row=row, column=get_col('PLANNED PICK UP DATE')))
                    revised_date = get_date_val(sheet.cell(row=row, column=get_col('Revised planned pick up date')))
                elif field == 'ATD Date':
                    planned_date = get_date_val(sheet.cell(row=row, column=col_orig_etd))
                    revised_date = get_date_val(sheet.cell(row=row, column=col_rev_etd))
                elif field == 'ATA Date':
                    planned_date = get_date_val(sheet.cell(row=row, column=col_orig_eta))
                    revised_date = get_date_val(sheet.cell(row=row, column=col_rev_eta))
                elif field == 'Actual delivery to door date':
                    planned_date = get_date_val(sheet.cell(row=row, column=col_plan_deliv))
                    revised_date = get_date_val(sheet.cell(row=row, column=col_rev_deliv))
                elif 'Date' in field:
                    eta = get_date_val(sheet.cell(row=row, column=col_orig_eta))
                    if eta and today > eta:
                        field_overdue = True
                else:
                    atd = get_date_val(sheet.cell(row=row, column=col_orig_etd))
                    if atd and today > atd:
                        field_overdue = True

                if (planned_date or revised_date) and not field_overdue:
                    field_overdue = is_overdue(planned_date, revised_date, val, today)
                
                if field_overdue:
                    if field in HIGHLIGHT_FIELDS:
                        overdue_highlight_fields.append(field)
                        cell.fill = YELLOW_FILL
                        if field not in DELIVERY_FIELDS:
                            is_delivery_only_overdue = False
                    else:
                        cell.fill = CLEAR_FILL
                else:
                    cell.fill = CLEAR_FILL
            else:
                cell.fill = CLEAR_FILL

        # 4. MILESTONE STATUS
        ms_cell = sheet.cell(row=row, column=col_ms_status)
        if overdue_highlight_fields:
            ms_cell.value = "Missing: " + ", ".join(overdue_highlight_fields)
        else:
            ms_cell.value = ""

        # 5. BOL COLOR LOGIC
        bol_cell = sheet.cell(row=row, column=col_bol)
        if not overdue_highlight_fields:
            bol_cell.font = BLACK_FONT
        else:
            origin_reg = str(sheet.cell(row=row, column=col_origin_reg).value).strip().upper()
            dest_reg = str(sheet.cell(row=row, column=col_dest_reg).value).strip().upper()
            origin_cntry = str(sheet.cell(row=row, column=col_origin_cntry).value).strip().upper()
            
                        # Post-Arrival (Delays after ATA)
            if is_delivery_only_overdue:
                if dest_reg == 'NO. AMERICA':
                    bol_cell.font = BLUE_FONT
                elif dest_reg in ['EUROPE', 'EMEA', 'MEDITERRANEAN']:
                    bol_cell.font = ORANGE_FONT
                elif dest_reg == 'ASIA':
                    bol_cell.font = RED_FONT
                else:
                    bol_cell.font = BLACK_FONT
            
            # Pre-Arrival (Origin and in-transit delays)
            else:
                if origin_cntry == 'GR' and dest_reg == 'NO. AMERICA':
                    bol_cell.font = BLUE_FONT
                elif origin_reg == 'ASIA':
                    bol_cell.font = RED_FONT
                elif origin_reg == 'NO. AMERICA':
                    bol_cell.font = BLUE_FONT
                elif origin_reg in ['EUROPE', 'EMEA', 'MEDITERRANEAN']:
                    bol_cell.font = ORANGE_FONT
                else:
                    bol_cell.font = BLACK_FONT
                
    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print("Complete!")

if __name__ == '__main__':
    main()

